#!/usr/bin/env python3
import sys
import os

# Add support for reading Excel files
try:
    import openpyxl
except ImportError:
    print("openpyxl not installed, trying alternative method...")
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl

from collections import Counter

file_path = r'C:\Users\Admin\V2Project\VJNT Class Managment\WebContent\Weekdays\UDISE वरील विद्यार्थी माहिती (Responses).xlsx'
output_path = r'C:\Users\Admin\V2Project\VJNT Class Managment\duplicate_udise_report.txt'

print(f"Reading file: {file_path}")
print(f"File exists: {os.path.exists(file_path)}")

try:
    # Load workbook
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    
    print(f"Sheet name: {ws.title}")
    print(f"Max row: {ws.max_row}")
    print(f"Max column: {ws.max_column}")
    print()
    
    # Find UDISE column
    udise_col = None
    header_row = 1
    
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(header_row, col)
        if cell.value and 'UDISE' in str(cell.value):
            udise_col = col
            print(f"Found UDISE column at column {col}")
            print(f"Header: {cell.value}")
            print()
            break
    
    if not udise_col:
        print("UDISE column not found! Available headers:")
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(header_row, col)
            print(f"  Column {col}: {cell.value}")
        
        with open(output_path, 'w', encoding='utf-8') as f:
            f.write("ERROR: UDISE column not found!\n")
        sys.exit(1)
    
    # Extract UDISE numbers
    udise_list = []
    for row in range(header_row + 1, ws.max_row + 1):
        cell = ws.cell(row, udise_col)
        if cell.value:
            udise_list.append(str(cell.value).strip())
    
    print(f"Total UDISE entries: {len(udise_list)}")
    print(f"Unique UDISE numbers: {len(set(udise_list))}")
    print()
    
    # Find duplicates
    counter = Counter(udise_list)
    duplicates = {udise: count for udise, count in counter.items() if count > 1}
    
    # Sort by count (descending)
    sorted_dups = sorted(duplicates.items(), key=lambda x: x[1], reverse=True)
    
    # Prepare output
    output_lines = []
    output_lines.append("DUPLICATE UDISE NUMBERS REPORT")
    output_lines.append("=" * 70)
    output_lines.append(f"Generated: {__import__('datetime').datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    output_lines.append(f"File: {file_path}")
    output_lines.append("")
    output_lines.append(f"Total UDISE entries: {len(udise_list)}")
    output_lines.append(f"Unique UDISE numbers: {len(set(udise_list))}")
    output_lines.append(f"Duplicate UDISE numbers found: {len(sorted_dups)}")
    output_lines.append("")
    output_lines.append("=" * 70)
    output_lines.append("DUPLICATE UDISE NUMBERS:")
    output_lines.append("=" * 70)
    output_lines.append("")
    
    if sorted_dups:
        output_lines.append("UDISE Number" + " " * 25 + "| Count")
        output_lines.append("-" * 70)
        for udise, count in sorted_dups:
            line = str(udise).ljust(35) + f"| {count}"
            output_lines.append(line)
        output_lines.append("")
        output_lines.append("=" * 70)
        output_lines.append("SUMMARY")
        output_lines.append("=" * 70)
        output_lines.append(f"Total duplicate UDISE numbers: {len(sorted_dups)}")
        total_dup_entries = sum(count for _, count in sorted_dups)
        output_lines.append(f"Total duplicate entries: {total_dup_entries}")
    else:
        output_lines.append("No duplicate UDISE numbers found!")
    
    output_lines.append("=" * 70)
    
    # Print to console
    output_text = "\n".join(output_lines)
    print(output_text)
    
    # Write to file
    with open(output_path, 'w', encoding='utf-8') as f:
        f.write(output_text)
    
    print(f"\n\nResults saved to: {output_path}")
    
except Exception as e:
    print(f"ERROR: {str(e)}")
    import traceback
    traceback.print_exc()
    
    with open(output_path, 'w', encoding='utf-8') as f:
        f.write(f"ERROR: {str(e)}\n")
    sys.exit(1)
